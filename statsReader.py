import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.utils import get_column_letter
import numpy as np
import csv
import os

# TODO: Timeline? TODO: differentiate quarter change: when T+(previous) > T+(current), that means there's a quarter 
#  change. To double check, if time difference is bigger than 15 min, quarter change 

MAX_ARRAY_ROWS = 10

# NEVER FORGET OUR BEST PLAYER: MR. PLACEHOLDER, HIS JERSEY NUMBER IS XX
headingResult = np.empty([10, 3], 'U50')
# expected structure: [STRING, Time, Time]
# note: https://stackoverflow.com/questions/55377213/numpy-taking-only-first-character-of-string
# type str only takes the first character of string
# use U + length instead
selfScoreResult = np.empty([MAX_ARRAY_ROWS, 5], 'U50')
opponentScoreResult = np.empty([MAX_ARRAY_ROWS, 5], 'U50')
# expected structure: [Main, Assist1, Assist2, Time Happened, Quarter T+]
selfSaveResult = np.empty([MAX_ARRAY_ROWS, 3], 'U50')
opponentSaveResult = np.empty([MAX_ARRAY_ROWS, 3], 'U50')
# expected structure: [Goalie, Time Happened, Quarter T+]
selfAttemptResult = np.empty([MAX_ARRAY_ROWS, 3], 'U50')
opponentAttemptResult = np.empty([MAX_ARRAY_ROWS, 3], 'U50')
# expected structure: [Main, Time Happened, Quarter T+]
selfGroundballResult = np.empty([MAX_ARRAY_ROWS, 3], 'U50')
opponentGroundballResult = np.empty([MAX_ARRAY_ROWS, 3], 'U50')
# expected structure: [Who Picked-up, Time Happened, Quarter T+]
selfTurnoverResult = np.empty([MAX_ARRAY_ROWS, 4], 'U50')
opponentTurnoverResult = np.empty([MAX_ARRAY_ROWS, 4], 'U50')
# expected structure：[Who Caused, Who Dropped, Time Happened, Quarter T+]
selfFaceoffResult = np.empty([MAX_ARRAY_ROWS, 4], 'U50')
opponentFaceoffResult = np.empty([MAX_ARRAY_ROWS, 4], 'U50')
# expected structure：[Who Won, Who Lost, Time Happened, Quarter T+]
selfPenaltyResult = np.empty([MAX_ARRAY_ROWS, 5], 'U50')
opponentPenaltyResult = np.empty([MAX_ARRAY_ROWS, 5], 'U50')
# expected structure: [Main, Type, Length, Time Happened, Quarter T+]
headingList = []  # storing info about game heading
selfList = []  # storing info about your team
opponentList = []  # storing info about the opponent team
selfScoreList = []
opponentScoreList = []
selfAttemptList = []
opponentAttemptList = []
selfSaveList = []
opponentSaveList = []
selfGroundballList = []
opponentGroundballList = []
selfTurnoverList = []
opponentTurnoverList = []
selfPenaltyList = []
opponentPenaltyList = []
selfFaceoffList = []
opponentFaceoffList = []

last = 0

with open('record.txt') as f:
    lines = [line.rstrip() for line in f]  # storing everything
f.close()

for self in lines:
    if self[0] != '#' and self[0] != '*':
        selfList.append(self)
    elif self[0] != '*':
        opponentList.append(self.replace('#', ''))


def get_heading():
    linesLen = len(lines)
    i = 0
    for x in range(linesLen):
        current = lines[i]
        try:
            if current[0] == 'i':
                removed = lines.pop(i)
                removed = removed[1:]
                headingList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    headingList.pop(0)
    headingList.pop(0)
    # i, x = 0, 0  # reset var i, x

    for x in range(len(headingList)):
        split = headingList[x].split(',')
        headingResult[x][0] = split[0]
        headingResult[x][1] = split[1]
        headingResult[x][2] = split[2]


def get_self_score():
    linesLen = len(selfList)
    i = 0
    for x in range(linesLen):
        current = selfList[i]
        try:
            if current[0] == 's':
                removed = selfList.pop(i)
                removed = removed[1:]
                selfScoreList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(selfScoreList)
    # i, x = 0, 0  # reset var i, x

    for x in range(len(selfScoreList)):
        numbers = []
        split = selfScoreList[x].split(',')

        while (len(split[0]) <= 6):
            split[0] = str(split[0]) + "xx"

        for i in range(len(split[0])):
            numbers.append(split[0][i])

        selfScoreResult[x][0] = str(numbers[0]) + str(numbers[1])
        selfScoreResult[x][1] = str(numbers[2]) + str(numbers[3])
        selfScoreResult[x][2] = str(numbers[4]) + str(numbers[5])
        selfScoreResult[x][3] = split[1]
        selfScoreResult[x][4] = split[2]


def get_opponent_score():
    linesLen = len(opponentList)
    i = 0
    for x in range(linesLen):
        current = opponentList[i]
        try:
            if current[0] == 's':
                removed = opponentList.pop(i)
                removed = removed[1:]
                opponentScoreList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(opponentScoreList)
    # i, x = 0, 0  # reset var i, x

    for x in range(len(opponentScoreList)):
        numbers = []
        split = opponentScoreList[x].split(',')

        while len(split[0]) <= 6:
            split[0] = str(split[0]) + "xx"

        for i in range(len(split[0])):
            numbers.append(split[0][i])

        opponentScoreResult[x][0] = str(numbers[0]) + str(numbers[1])
        opponentScoreResult[x][1] = str(numbers[2]) + str(numbers[3])
        opponentScoreResult[x][2] = str(numbers[4]) + str(numbers[5])
        opponentScoreResult[x][3] = split[1]
        opponentScoreResult[x][4] = split[2]


def get_self_attempt():
    linesLen = len(selfList)
    i = 0
    for x in range(linesLen):
        current = selfList[i]
        try:
            if current[0] == 'a':
                removed = selfList.pop(i)
                removed = removed[1:]
                selfAttemptList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(selfScoreList)
    # i, x = 0, 0  # reset var i, x
    for x in range(len(selfAttemptList)):
        split = selfAttemptList[x].split(',')
        selfAttemptResult[x][0] = split[0]
        selfAttemptResult[x][1] = split[1]
        selfAttemptResult[x][2] = split[2]


def get_opponent_attempt():
    linesLen = len(opponentList)
    i = 0
    for x in range(linesLen):
        current = opponentList[i]
        try:
            if current[0] == 'a':
                removed = opponentList.pop(i)
                removed = removed[1:]
                opponentAttemptList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(opponentScoreList)
    # i, x = 0, 0  # reset var i, x
    for x in range(len(opponentAttemptList)):
        split = opponentAttemptList[x].split(',')
        opponentAttemptResult[x][0] = split[0]
        opponentAttemptResult[x][1] = split[1]
        opponentAttemptResult[x][2] = split[2]
        

def get_self_save():
    linesLen = len(selfList)
    i = 0
    for x in range(linesLen):
        current = selfList[i]
        try:
            if current[0] == 'g':
                removed = selfList.pop(i)
                removed = removed[1:]
                selfSaveList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(selfScoreList)
    # i, x = 0, 0  # reset var i, x
    for x in range(len(selfSaveList)):
        split = selfSaveList[x].split(',')
        selfSaveResult[x][0] = split[0]
        selfSaveResult[x][1] = split[1]
        selfSaveResult[x][2] = split[2]


def get_opponent_save():
    linesLen = len(opponentList)
    i = 0
    for x in range(linesLen):
        current = opponentList[i]
        try:
            if current[0] == 'g':
                removed = opponentList.pop(i)
                removed = removed[1:]
                opponentSaveList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(opponentScoreList)
    # i, x = 0, 0  # reset var i, x
    for x in range(len(opponentSaveList)):
        split = opponentSaveList[x].split(',')
        opponentSaveResult[x][0] = split[0]
        opponentSaveResult[x][1] = split[1]
        opponentSaveResult[x][2] = split[2]


def get_self_groundball():
    linesLen = len(selfList)
    i = 0
    for x in range(linesLen):
        current = selfList[i]
        try:
            if current[0] == 'b':
                removed = selfList.pop(i)
                removed = removed[1:]
                selfGroundballList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(selfScoreList)
    # i, x = 0, 0  # reset var i, x
    for x in range(len(selfGroundballList)):
        split = selfGroundballList[x].split(',')
        if split[0] == '':
            split[0] = "xx"
        selfGroundballResult[x][0] = split[0]
        selfGroundballResult[x][1] = split[1]
        selfGroundballResult[x][2] = split[2]


def get_opponent_groundball():
    linesLen = len(opponentList)
    i = 0
    for x in range(linesLen):
        current = opponentList[i]
        try:
            if current[0] == 'b':
                removed = opponentList.pop(i)
                removed = removed[1:]
                opponentGroundballList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(opponentScoreList)
    # i, x = 0, 0  # reset var i, x
    for x in range(len(opponentGroundballList)):
        split = opponentGroundballList[x].split(',')
        if split[0] == '':
            split[0] = "xx"
        opponentGroundballResult[x][0] = split[0]
        opponentGroundballResult[x][1] = split[1]
        opponentGroundballResult[x][2] = split[2]


def get_self_penalty():
    linesLen = len(selfList)
    i = 0
    for x in range(linesLen):
        current = selfList[i]
        try:
            if current[0] == 'p':
                removed = selfList.pop(i)
                removed = removed[1:]
                selfPenaltyList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(selfScoreList)
    # i, x = 0, 0  # reset var i, x
    # expected structure: [Main, Type, Length, Time Happened, Quarter T+]
    # input: p02t2
    for x in range(len(selfPenaltyList)):
        split = selfPenaltyList[x].split(',')
        selfPenaltyResult[x][0] = split[0][0] + split[0][1]
        selfPenaltyResult[x][1] = split[0][2]
        selfPenaltyResult[x][2] = split[0][3]
        selfPenaltyResult[x][3] = split[1]
        selfPenaltyResult[x][4] = split[2]


def get_opponent_penalty():
    linesLen = len(opponentList)
    i = 0
    for x in range(linesLen):
        current = opponentList[i]
        try:
            if current[0] == 'p':
                removed = opponentList.pop(i)
                removed = removed[1:]
                opponentPenaltyList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(opponentScoreList)
    # i, x = 0, 0  # reset var i, x
    # expected structure: [Main, Type, Length, Time Happened, Quarter T+]
    # input: p02t2
    for x in range(len(opponentPenaltyList)):
        split = opponentPenaltyList[x].split(',')
        opponentPenaltyResult[x][0] = split[0][0] + split[0][1]
        opponentPenaltyResult[x][1] = split[0][2]
        opponentPenaltyResult[x][2] = split[0][3]
        opponentPenaltyResult[x][3] = split[1]
        opponentPenaltyResult[x][4] = split[2]


def get_self_turnover():
    linesLen = len(selfList)
    i = 0
    for x in range(linesLen):
        current = selfList[i]
        try:
            if current[0] == 't':
                removed = selfList.pop(i)
                removed = removed[1:]
                selfTurnoverList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(selfTurnoverList)
    # i, x = 0, 0  # reset var i, x

    for x in range(len(selfTurnoverList)):
        numbers = []
        split = selfTurnoverList[x].split(',')

        while len(split[0]) <= 4:
            split[0] = str(split[0]) + "xx"

        for i in range(len(split[0])):
            numbers.append(split[0][i])
        # expected structure：[Who Caused, Who Dropped, Time Happened, Quarter T+]
        selfTurnoverResult[x][0] = str(numbers[0]) + str(numbers[1])
        selfTurnoverResult[x][1] = str(numbers[2]) + str(numbers[3])
        selfTurnoverResult[x][2] = split[1]
        selfTurnoverResult[x][3] = split[2]


def get_opponent_turnover():
    linesLen = len(opponentList)
    i = 0
    for x in range(linesLen):
        current = opponentList[i]
        try:
            if current[0] == 't':
                removed = opponentList.pop(i)
                removed = removed[1:]
                opponentTurnoverList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(opponentTurnoverList)
    # i, x = 0, 0  # reset var i, x

    for x in range(len(opponentTurnoverList)):
        numbers = []
        split = opponentTurnoverList[x].split(',')

        while len(split[0]) <= 4:
            split[0] = str(split[0]) + "xx"

        for i in range(len(split[0])):
            numbers.append(split[0][i])
        # expected structure：[Who Caused, Who Dropped, Time Happened, Quarter T+]
        opponentTurnoverResult[x][0] = str(numbers[0]) + str(numbers[1])
        opponentTurnoverResult[x][1] = str(numbers[2]) + str(numbers[3])
        opponentTurnoverResult[x][2] = split[1]
        opponentTurnoverResult[x][3] = split[2]


def get_self_faceoff():
    linesLen = len(selfList)
    i = 0
    for x in range(linesLen):
        current = selfList[i]
        try:
            if current[0] == 'f':
                removed = selfList.pop(i)
                removed = removed[1:]
                selfFaceoffList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(selfFaceoffList)
    # i, x = 0, 0  # reset var i, x

    for x in range(len(selfFaceoffList)):
        numbers = []
        split = selfFaceoffList[x].split(',')

        while len(split[0]) <= 4:
            split[0] = str(split[0]) + "xx"

        for i in range(len(split[0])):
            numbers.append(split[0][i])
        # expected structure：[Who Caused, Who Dropped, Time Happened, Quarter T+]
        selfFaceoffResult[x][0] = str(numbers[0]) + str(numbers[1])
        selfFaceoffResult[x][1] = str(numbers[2]) + str(numbers[3])
        selfFaceoffResult[x][2] = split[1]
        selfFaceoffResult[x][3] = split[2]


def get_opponent_faceoff():
    linesLen = len(opponentList)
    i = 0
    for x in range(linesLen):
        current = opponentList[i]
        try:
            if current[0] == 'f':
                removed = opponentList.pop(i)
                removed = removed[1:]
                opponentFaceoffList.append(removed)
                i -= 1
        except IndexError:
            print("Current line is blank")
        i += 1
    # print(opponentFaceoffList)
    # i, x = 0, 0  # reset var i, x

    for x in range(len(opponentFaceoffList)):
        numbers = []
        split = opponentFaceoffList[x].split(',')

        while len(split[0]) <= 4:
            split[0] = str(split[0]) + "xx"

        for i in range(len(split[0])):
            numbers.append(split[0][i])
        # expected structure：[Who Caused, Who Dropped, Time Happened, Quarter T+]
        opponentFaceoffResult[x][0] = str(numbers[0]) + str(numbers[1])
        opponentFaceoffResult[x][1] = str(numbers[2]) + str(numbers[3])
        opponentFaceoffResult[x][2] = split[1]
        opponentFaceoffResult[x][3] = split[2]


# Testing Area


get_heading()
print("HEADING")
print(headingResult)

get_self_score()
print("SELF SCORE")
print(selfScoreResult)

get_opponent_score()
print("OPPONENT SCORE")
print(opponentScoreResult)

get_self_attempt()
print("SELF ATTEMPT")
print(selfAttemptResult)

get_opponent_attempt()
print("OPPONENT ATTEMPT")
print(opponentAttemptResult)

get_self_save()
print("SELF SAVE")
print(selfSaveResult)

get_opponent_save()
print("OPPONENT SAVE")
print(opponentSaveResult)

get_self_groundball()
print("SELF GROUNDBALL")
print(selfGroundballResult)

get_opponent_groundball()
print("OPPONENT GROUNDBALL")
print(opponentGroundballResult)

get_self_penalty()
print("SELF PENALTY")
print(selfPenaltyResult)

get_opponent_penalty()
print("OPPONENT PENALTY")
print(opponentPenaltyResult)

get_self_turnover()
print("SELF TURNOVER")
print(selfTurnoverResult)

get_opponent_turnover()
print("OPPONENT TURNOVER")
print(opponentTurnoverResult)

get_self_faceoff()
print("SELF FACEOFF")
print(selfFaceoffResult)

get_opponent_faceoff()
print("OPPONENT FACEOFF")
print(opponentFaceoffResult)

# Finished putting all the data in arrays
# Write in .xlsx
# goalieFill = PatternFill(patternType='solid', fill_type='solid', bgColor=colors.RED)
#  = PatternFill(start_color='FFCCFFFF', end_color=
titleFill = PatternFill("solid", fgColor="FFA500")
goalieFill = PatternFill("solid", fgColor="87CEFA")
offenseFill = PatternFill("solid", fgColor="FFFACD")
midiFill = PatternFill("solid", fgColor="C0D9AF")
defenseFill = PatternFill("solid", fgColor="F0F8FF")
foFill = PatternFill("solid", fgColor="D8BFD8")

data_path = 'selfPlayer.csv'
with open(data_path) as f:
    reader = csv.reader(f, delimiter=',')
    selfPlayers = np.array(list(reader)).astype('U50')

if len(selfPlayers) > MAX_ARRAY_ROWS:
    MAX_ARRAY_ROWS = len(selfPlayers) + 2

print(selfPlayers)

wb = Workbook()
filename = headingResult[0][0] + ' vs ' + headingResult[3][0] + '.xlsx'
ws1 = wb.active
ws1.title = "Scorebook"
heading = ['Self', 'Coach', 'Record', 'Opponent', 'Coach', 'Record']
ws1.merge_cells('A1:C1')
ws1['A1'] = "Team Information"
ws1['A1'].fill = titleFill
for x in range(6):
    ws1['A' + str(x + 2)] = heading[x]
    ws1['A' + str(x + 2)].fill = titleFill
    ws1.merge_cells('B' + str(x + 2) + ':' + 'C' + str(x + 2))
    ws1['B' + str(x + 2)] = headingResult[x][0]

ws1.merge_cells('A10:P10')
ws1['A10'] = ('TEAM: ' + headingResult[0][0])
ws1['A10'].alignment = Alignment(horizontal='center')
ws1['A11'] = 'POS'
ws1['A11'].fill = titleFill
ws1['B11'] = 'NO'
ws1['B11'].fill = titleFill
ws1['C11'] = 'NAME'
ws1['C11'].fill = titleFill
ws1['G11'] = 'ATPTFAIL'
ws1['G11'].fill = titleFill
ws1['H11'] = 'GOAL'
ws1['H11'].fill = titleFill
ws1['I11'] = 'ASSIST'
ws1['I11'].fill = titleFill
ws1['J11'] = 'GRBALL'
ws1['J11'].fill = titleFill
ws1['K11'] = 'TURNOVER'
ws1['K11'].fill = titleFill
ws1['L11'] = 'PENALTY'
ws1['L11'].fill = titleFill
ws1['M11'] = 'GOSAVE'
ws1['M11'].fill = titleFill
ws1['N11'] = 'GOMISS'
ws1['N11'].fill = titleFill
ws1['O11'] = 'FOWIN'
ws1['O11'].fill = titleFill
ws1['P11'] = 'FOLOSE'
ws1['P11'].fill = titleFill
ws1.merge_cells('C11:F11')
ws1['A10'].fill = titleFill


def find_line(number):
    for x in range(1, 100):
        if ws1.cell(row=x, column=2).value == number:
            print('LINE ' + str(x))
            return x
    for x in range(1, 100):
        if ws1.cell(row=x, column=2).value == 'xx':
            print('LINE ' + str(x))
            return x


def find_line_ignore_xx(number):
    for x in range(1, 100):
        if ws1.cell(row=x, column=2).value == number:
            return x
    return False


for x in range(MAX_ARRAY_ROWS):
    ws1.merge_cells('C' + str(x + 12) + ':' + 'F' + str(x + 12))
    ws1['A' + str(x + 12)] = selfPlayers[x][0]
    last = x + 13
    if selfPlayers[x][0] == 'G':
        for y in range(1, 17):
            ws1.cell(row=(x + 12), column=y).fill = goalieFill
        for z in range(7, 17):
            try:
                ws1.cell(row=(x + 12), column=z).value = 0
            except AttributeError:
                None
    elif selfPlayers[x][0] == 'A':
        for y in range(1, 17):
            ws1.cell(row=(x + 12), column=y).fill = offenseFill
        for z in range(7, 17):
            try:
                ws1.cell(row=(x + 12), column=z).value = 0
            except AttributeError:
                None
    elif selfPlayers[x][0] == 'M':
        for y in range(1, 17):
            ws1.cell(row=(x + 12), column=y).fill = midiFill
        for z in range(7, 17):
            try:
                ws1.cell(row=(x + 12), column=z).value = 0
            except AttributeError:
                None
    elif selfPlayers[x][0] == 'D':
        for y in range(1, 17):
            ws1.cell(row=(x + 12), column=y).fill = defenseFill
        for z in range(7, 17):
            try:
                ws1.cell(row=(x + 12), column=z).value = 0
            except AttributeError:
                None
    elif selfPlayers[x][0] == 'FM':
        for y in range(1, 17):
            ws1.cell(row=(x + 12), column=y).fill = foFill
        for z in range(7, 17):
            try:
                ws1.cell(row=(x + 12), column=z).value = 0
            except AttributeError:
                None
    # TODO: split midi to amidi an dmidi?
    ws1['B' + str(x + 12)] = selfPlayers[x][1]
    ws1['C' + str(x + 12)] = selfPlayers[x][2]

for x in range(7, 17):
    ws1.cell(row=last, column=x).value = 0
ws1['B' + str(last)] = 'xx'
ws1.merge_cells('C' + str(last) + ':' + 'F' + str(last))
ws1.cell(row=last, column=3).value = 'PLACEHOLDER'
# placeholder player

# score & assist
i = 0
for x in range(12, len(selfScoreResult) + 12):
    if selfScoreResult[i][0] != 'xx' and selfScoreResult[i][0] != '':
        row = find_line(selfScoreResult[i][0])
        print(row)
        ws1['H' + str(row)] = int(ws1['H' + str(row)].value) + 1
    if selfScoreResult[i][1] != 'xx' and selfScoreResult[i][0] != '':
        row = find_line(selfScoreResult[i][0])
        ws1['I' + str(row)] = int(ws1['I' + str(row)].value) + 1
    if selfScoreResult[i][2] != 'xx' and selfScoreResult[i][0] != '':
        row = find_line(selfScoreResult[i][0])
        ws1['I' + str(row)] = int(ws1['I' + str(row)].value) + 1
    i += 1
i = 0

# attempt
i = 0
for x in range(12, len(selfAttemptResult) + 12):
    if selfAttemptResult[i][0] != 'xx' and selfAttemptResult[i][0] != '':
        row = find_line(selfAttemptResult[i][0])
        print(row)
        ws1['G' + str(row)] = int(ws1['G' + str(row)].value) + 1
    i += 1
i = 0
wb.save(filename=filename)
os.startfile(filename, 'open')
